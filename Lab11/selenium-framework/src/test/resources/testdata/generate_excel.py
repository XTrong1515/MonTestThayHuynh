import openpyxl

wb = openpyxl.Workbook()
ws_smoke = wb.active
ws_smoke.title = "SmokeCases"

ws_smoke.append(["username", "password", "expected_url", "description"])
ws_smoke.append(["standard_user", "secret_sauce", "inventory", "Đăng nhập hợp lệ"])
ws_smoke.append(["problem_user", "secret_sauce", "inventory", "Đăng nhập với problem_user"])
ws_smoke.append(["performance_glitch_user", "secret_sauce", "inventory", "Đăng nhập với user chậm"])

ws_negative = wb.create_sheet("NegativeCases")
ws_negative.append(["username", "password", "expected_error", "description"])
ws_negative.append(["locked_out_user", "secret_sauce", "Epic sadface: Sorry, this user has been locked out.", "User bị khoá"])
ws_negative.append(["standard_user", "wrong", "Epic sadface: Username and password do not match any user in this service", "Sai mật khẩu"])
ws_negative.append(["", "secret_sauce", "Epic sadface: Username is required", "Username rỗng"])
ws_negative.append(["standard_user", "", "Epic sadface: Password is required", "Password rỗng"])
ws_negative.append(["wrong_user", "wrong_pass", "Epic sadface: Username and password do not match any user in this service", "Sai tải khoản và mật khẩu"])

ws_boundary = wb.create_sheet("BoundaryCases")
ws_boundary.append(["username", "password", "expected_error", "description"])
ws_boundary.append(["a" * 100, "secret_sauce", "Epic sadface: Username and password do not match any user in this service", "Tên đăng nhập rất dài"])
ws_boundary.append(["<script>alert(1)</script>", "secret_sauce", "Epic sadface: Username and password do not match any user in this service", "XSS injection attempt"])
ws_boundary.append(["' OR 1=1 --", "secret_sauce", "Epic sadface: Username and password do not match any user in this service", "SQL injection attempt"])
ws_boundary.append(["standard_user", "a" * 100, "Epic sadface: Username and password do not match any user in this service", "Mật khẩu rất dài"])

wb.save("login_data.xlsx")
print("Tạo thành công login_data.xlsx")
